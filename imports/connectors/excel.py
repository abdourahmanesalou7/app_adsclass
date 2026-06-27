"""
Connecteur Excel (.xlsx / .xlsm / .xls via openpyxl).
"""
from openpyxl import load_workbook
from .base import BaseConnector
from ..security import sanitize_cell, enforce_row_limit, enforce_column_limit


class ExcelConnector(BaseConnector):
    source_type = 'excel'

    def extract(self, filepath, sheet_name=None, header_row=1, **kwargs):
        wb = load_workbook(filepath, read_only=True, data_only=True)
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            wb.close()
            return {'headers': [], 'rows': []}

        header_idx = max(0, header_row - 1)
        raw_headers = all_rows[header_idx] if header_idx < len(all_rows) else []
        headers = self._normalize_headers(raw_headers)
        enforce_column_limit(headers)

        data_rows = []
        for row in all_rows[header_idx + 1:]:
            if row is None or all(v is None or (isinstance(v, str) and not v.strip()) for v in row):
                continue
            record = {}
            for i, header in enumerate(headers):
                val = row[i] if i < len(row) else None
                record[header] = sanitize_cell(val) if isinstance(val, str) else val
            data_rows.append(record)

        enforce_row_limit(len(data_rows))
        wb.close()
        return {
            'headers': headers,
            'rows': data_rows,
            'meta': {'sheet': ws.title, 'sheets_available': wb.sheetnames if hasattr(wb, 'sheetnames') else []},
        }
